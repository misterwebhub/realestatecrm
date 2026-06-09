import os
from openpyxl import load_workbook
p = r"c:\xampp\htdocs\realestate\arazis-map\extracted_tables\tables_59-46.xlsx"
print('exists', os.path.exists(p))
print('size', os.path.getsize(p) if os.path.exists(p) else None)
try:
    wb = load_workbook(p)
    print('sheets:', wb.sheetnames)
except Exception as e:
    print('error loading workbook:', e)
