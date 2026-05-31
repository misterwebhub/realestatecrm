import os
import sys
import subprocess

# Install requirements if missing
reqs = ["pdfplumber", "pandas", "openpyxl"]
for pkg in reqs:
    try:
        __import__(pkg)
    except Exception:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

import pdfplumber
import pandas as pd

def main():
    pdf_path = os.path.join(os.path.dirname(__file__), "319.pdf")
    out_dir = os.path.join(os.path.dirname(__file__), "extracted_tables")
    os.makedirs(out_dir, exist_ok=True)

    start_page = 46
    end_page = 59

    writer_path = os.path.join(out_dir, f"tables_{end_page}-{start_page}.xlsx")

    logs = []
    writer = None

    with pdfplumber.open(pdf_path) as pdf:
        for pnum in range(start_page, end_page+1):
            page_index = pnum - 1
            if page_index < 0 or page_index >= len(pdf.pages):
                logs.append(f"Page {pnum} out of range (pdf has {len(pdf.pages)} pages)")
                continue
            page = pdf.pages[page_index]
            tables = page.extract_tables()
            if not tables:
                logs.append(f"No tables found on page {pnum}")
                continue
            # create writer on first found table
            if writer is None:
                writer = pd.ExcelWriter(writer_path, engine="openpyxl")
            for ti, table in enumerate(tables, start=1):
                df = pd.DataFrame(table)
                if df.shape[0] > 1 and df.iloc[0].isnull().sum() < df.shape[1]:
                    df.columns = df.iloc[0]
                    df = df[1:].reset_index(drop=True)
                csv_name = f"page_{pnum}_table{ti}.csv"
                csv_path = os.path.join(out_dir, csv_name)
                df.to_csv(csv_path, index=False)
                sheet_name = f"p{pnum}_t{ti}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                logs.append(f"Extracted table {ti} from page {pnum}: saved {csv_name} and sheet {sheet_name}")

    if writer is not None:
        writer.close()
        # Reorder sheets descending by page
        from openpyxl import load_workbook
        wb = load_workbook(writer_path)

        def sheet_page_key(name):
            try:
                if name.startswith('p'):
                    parts = name[1:].split('_')
                    p = int(parts[0])
                    t = int(parts[1][1:]) if len(parts) > 1 and parts[1].startswith('t') else 0
                    return (-p, t)
            except Exception:
                pass
            return (0, 0)

        sheets = list(wb.sheetnames)
        sheets_sorted = sorted(sheets, key=sheet_page_key)
        wb._sheets = [wb[s] for s in sheets_sorted]
        wb.save(writer_path)

    log_path = os.path.join(out_dir, "extraction_log.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(logs) if logs else "No tables detected on pages 46-59")

    print("DONE")
    print("Output dir:", out_dir)
    print("Workbook:", writer_path if writer is not None else "(no workbook created)")
    print("Log:", log_path)

if __name__ == '__main__':
    main()
